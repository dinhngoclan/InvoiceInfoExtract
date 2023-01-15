# -*- coding: utf-8 -*-
"""
Created on Thu Apr 28 17:05:24 2022

@author: ngoclan
"""

#=======================Convert .xls to .xlsx========================
import os, glob
import win32com.client as win32
from win32com import client
import time
#time.sleep(5)
def convert_xls2xlsx(path):
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    #excel = client.Dispatch("Excel.Application")
    filenames = glob.glob(path + "\*.xls")
    for file in filenames:
        #os.path.abspath(r".\TCDKO-2204-08 120422 (Kumho-2309)-01 NYLON-BUSAN-XKD.XLS")
        print ("convert_xls2xlsx: ",file)
        wb = excel.Workbooks.Open(file)
        #time.sleep(0.5)
        wb.SaveAs(file+'x', FileFormat = 51)    #FileFormat = 51 is for .xlsx extension        
        wb.Close()                               #FileFormat = 56 is for .xls extension
        wb = None
    excel.Application.Quit()

convert_xls2xlsx(os.path.abspath(r".\InvoicesDirectory"))

#=======================Hide GridLine in .xlsx ========================
##OK OK
import openpyxl
from openpyxl.styles.alignment import Alignment
# Load workbook
import pyexcel
#pyexcel.save_book_as(file_name='.\TCDKO-2204-08 120422 (Kumho-2309)-01 NYLON-BUSAN-XKD.XLS', dest_file_name='.\TCDKO-2204-08 120422 (Kumho-2309)-01 NYLON-BUSAN-XKD.XLSX')
def hide_gridline_in_xlsx(path):
    filenames = glob.glob(path + "\*.xlsx")
    for file in filenames:
        wb = openpyxl.load_workbook(file)
        # Loop through all cells in all worksheets
        for sheet in wb.worksheets:
            sheet.sheet_view.showGridLines = False
            sheet.sheet_view.view='normal'
            sheet.paper_size_code = 1
          
            #sheet.alignment = Alignment(horizontal="center")
            #remove "///////////////////////////"
            for r in range(1,sheet.max_row+1):
                for c in range(1,sheet.max_column+1):
                    s = sheet.cell(r,c).value
                    try:
                        if s != None and "/////////////////////////////////////////////////////////////" in s: 
                            sheet.cell(r,c).value = ''
                    except:
                        pass
        # Save workbook
        wb.save(file)
        wb.close()
 
hide_gridline_in_xlsx(os.path.abspath(r".\InvoicesDirectory"))

# =============================================================================
# #======================Convert .xlsx to .png ========================
# #OK OK
# import excel2img, glob
# def xlsx2img(path):
#     filenames = glob.glob(path + "\*.xlsx")
#     for file in filenames:
#         excel2img.export_img(file, file+'.png', "INVOICE", None)
# 
# xlsx2img(os.path.abspath(r".\InvoicesDirectory"))
# =============================================================================

#======================Convert .xlsx to .pdf ========================
#from win32com import client
    
def xlsx2pdf(path):
    filenames = glob.glob(path + "\*.xlsx")
    xlApp = client.Dispatch("Excel.Application")
    #xlApp = win32.gencache.EnsureDispatch('Excel.Application')
    #xlApp.Visible = False
    for file in filenames:
        print("xlsx2pdf: ", file)
        books = xlApp.Workbooks.Open(file)
       
        try:
            wsInvoice = books.Worksheets['INVOICE']
            wsPackageList = books.Worksheets['PACKING LIST']
        except:
            try:
                wsInvoice = books.Worksheets['Invoice ']
                wsPackageList = books.Worksheets['PList']
            except:
                wsInvoice = books.Worksheets['INV CIP']
                wsPackageList = books.Worksheets['PACKING LIST']
        wsInvoice.EnableCalculation = True
        wsInvoice.Calculate()
        wsInvoice.PageSetup.Zoom = False
        wsInvoice.PageSetup.FitToPagesTall = 1
        wsInvoice.PageSetup.FitToPagesWide = 1
        
        wsPackageList.EnableCalculation = True
        wsPackageList.Calculate()
        wsPackageList.PageSetup.Zoom = False
        wsPackageList.PageSetup.FitToPagesTall = 1
        wsPackageList.PageSetup.FitToPagesWide = 1      
        books.Save()
        #time.sleep(5)
        #ws.Range("A1", "P100").HorizontalAlignment = 2  #align LEFT 
        #ws.Range("A1", "P80").HorizontalAlignment = 1 #align RIGHT
        try:
            wsInvoice.SaveAs(file+'INVOICE.pdf', FileFormat=57)
            wsPackageList.SaveAs(file+'PACKING LIST.pdf', FileFormat=57)
            books.Close(True)
        except Exception as e:
            print("Failed to convert xlsx to pdf")
            print(str(e))
        #finally:
            
    xlApp.Quit()
    


xlsx2pdf(os.path.abspath(r".\InvoicesDirectory"))


#======================Convert .pdf to .png ========================
from pdf2image import convert_from_path
filenames = glob.glob('.\InvoicesDirectory' + "\*.pdf")
for file in filenames:
    images = convert_from_path(file,poppler_path=r'.\poppler-22.01.0\Library\bin')
    images[0].save(file+'.png', "PNG")

# -*- coding: utf-8 -*-
"""
Created on Thu Apr 28 17:05:24 2022

@author: ngoclan
"""


import os
#os.system('pip install git+https://github.com/huggingface/transformers.git --upgrade')
#os.system('pip install pyyaml==5.1')
# workaround: install old version of pytorch since detectron2 hasn't released packages for pytorch 1.9 (issue: https://github.com/facebookresearch/detectron2/issues/3158)
#os.system('pip install torch==1.8.0+cu101 torchvision==0.9.0+cu101 -f https://download.pytorch.org/whl/torch_stable.html')

# install detectron2 that matches pytorch 1.8
# See https://detectron2.readthedocs.io/tutorials/install.html for instructions
#os.system('pip install -q detectron2 -f https://dl.fbaipublicfiles.com/detectron2/wheels/cu101/torch1.8/index.html')

## install PyTesseract
#os.system('pip install -q pytesseract')



import numpy as np
import pandas as pd
from transformers import LayoutLMv2FeatureExtractor, LayoutLMv2TokenizerFast, LayoutLMv2Tokenizer, LayoutLMv2ForTokenClassification
from PIL import Image, ImageDraw, ImageFont, ImageOps
import datefinder
from dateutil.parser import parse
from price_parser import Price


feature_extractor = LayoutLMv2FeatureExtractor.from_pretrained(r".\savedmodel")#("microsoft/layoutlmv2-base-uncased")
tokenizer = LayoutLMv2TokenizerFast.from_pretrained(r".\savedmodel")#("microsoft/layoutlmv2-base-uncased")
model = LayoutLMv2ForTokenClassification.from_pretrained(r".\savedmodel")

# define id2label, label2color
#labels = dataset.features['ner_tags'].feature.names
labels = ['O', 'B-HEADER', 'I-HEADER', 'B-QUESTION', 'I-QUESTION', 'B-ANSWER', 'I-ANSWER']
id2label = {v: k for v, k in enumerate(labels)}
label2color = {'question':'blue', 'answer':'green', 'header':'orange', 'other':'violet'}

def unnormalize_box(bbox, width, height):
     return [
         width * (bbox[0] / 1000),
         height * (bbox[1] / 1000),
         width * (bbox[2] / 1000),
         height * (bbox[3] / 1000),
     ]

def iob_to_label(label):
    label = label[2:]
    if not label:
      return 'other'
    return label

def shear(img, shear):
    shear = img.transform(img.size, Image.AFFINE, (1, shear, 0, 0, 1, 0))
    return shear

def process_image(model, image):
    width, height = image.size
    
    # get words, boxes
    encoding_feature_extractor = feature_extractor(image, return_tensors="pt")
    words, boxes = encoding_feature_extractor.words, encoding_feature_extractor.boxes

    # encode
    encoding = tokenizer(words, boxes=boxes, return_offsets_mapping=True, return_tensors="pt", truncation=True)
    offset_mapping = encoding.pop('offset_mapping')
    encoding["image"] = encoding_feature_extractor.pixel_values

    # forward pass
    outputs = model(**encoding)

    # get predictions
    predictions = outputs.logits.argmax(-1).squeeze().tolist()
    token_boxes = encoding.bbox.squeeze().tolist()

    # only keep non-subword predictions
    is_subword = np.array(offset_mapping.squeeze().tolist())[:,0] != 0
    true_predictions = [id2label[pred] for idx, pred in enumerate(predictions) if not is_subword[idx]]
    true_boxes = [unnormalize_box(box, width, height) for idx, box in enumerate(token_boxes) if not is_subword[idx]]

  
    origBoxes = []
    for box in boxes[0] : 
        origBoxes.append(unnormalize_box(box, width, height))
    
    pdWordsBoxes = pd.concat([pd.DataFrame({'words': words[0]}), pd.DataFrame({'origBoxes': origBoxes})], axis=1, join='inner')
    pdWordsBoxes['origBoxes_STR'] = pdWordsBoxes['origBoxes'].astype(str)
    pdPred = pd.DataFrame(zip(true_boxes, true_predictions),columns = ['true_boxes', 'true_predictions'])
    pdPred['true_boxes_STR'] = pdPred['true_boxes'].astype(str)
    finalPD = pd.merge(pdWordsBoxes, pdPred, how='left', left_on = 'origBoxes_STR', right_on = 'true_boxes_STR')
    finalPD = finalPD.dropna(subset = ['true_predictions'])
    #remove all linhtinh character
    finalPD = finalPD[~finalPD.words.isin(["&", ",", ".", ":"]) ]
    
    for index, row in finalPD.iterrows():
        try:
            is_date = False
            for match in datefinder.find_dates( row['words']):
                is_date = True
            if is_date and 'ANSWER' not in row['true_predictions']:
                finalPD.at[index, 'true_predictions'] = "I-ANSWER"
        except:
            pass
    
    
    # draw predictions over the image
    draw = ImageDraw.Draw(image)
    font = ImageFont.load_default()
    for prediction, box in zip(finalPD['true_predictions'], finalPD['true_boxes']):
        #print(prediction)
        predicted_label = iob_to_label(prediction).lower()
        draw.rectangle(box, outline=label2color[predicted_label])
        draw.text((box[0]+10, box[1]-10), text=predicted_label, fill=label2color[predicted_label], font=font)
  
    return image, finalPD


import glob
def do_extract_from_png(image, file, model, graph):
    width, height = image.size
    #===================== Use ML to classify word ========================
    image, table = process_image(model, image)
    image.save(file+'.jpg')
    #================ Use Graph to math Question-Answer ===================
    tx = graph.begin()
    tx.evaluate('''  MATCH (n) DETACH DELETE n  ''')
    graph.commit(tx)
    #CREATE NODES
    tx = graph.begin()
    for index, row in table.iterrows():
        try:
            if row['true_predictions'] == 'O':
                tx.evaluate('''
                   MERGE (o:OTHER  {word:$label2, x0:$x0, y0:$y0, x1:$x1, y1:$y1})
                   ''', parameters = {'label2': row['words'], 'x0':row['origBoxes'][0], 'y0':row['origBoxes'][1]
                                                           , 'x1':row['origBoxes'][2], 'y1':row['origBoxes'][3]})
            elif 'QUESTION' in row['true_predictions']:
                tx.evaluate('''
                   MERGE (q:QUESTION {word:$label2, x0:$x0, y0:$y0, x1:$x1, y1:$y1})
                   ''', parameters = {'label2': row['words'], 'x0':row['origBoxes'][0], 'y0':row['origBoxes'][1]
                                                           , 'x1':row['origBoxes'][2], 'y1':row['origBoxes'][3]})
            elif 'ANSWER' in row['true_predictions']:
                tx.evaluate('''
                   MERGE (a:ANSWER {word:$label2, x0:$x0, y0:$y0, x1:$x1, y1:$y1})
                   ''', parameters = {'label2': row['words'], 'x0':row['origBoxes'][0], 'y0':row['origBoxes'][1]
                                                           , 'x1':row['origBoxes'][2], 'y1':row['origBoxes'][3]})
        except:
            print(row['true_predictions'])
        
    graph.commit(tx)
    
    #CREATE RELATIONSHIPs
    tx = graph.begin()
    tx.evaluate('''
                match (a:QUESTION), (b:QUESTION)
                where a.x0 - b.x0 < 0 and  a.x1 - b.x0 > -20
                     and abs(a.y0 - b.y0) < 10 
                create (a)-[:NEXTWORD {x:abs(a.x1-b.x0)}]->(b);
                ''')
                
    tx.evaluate('''
                match (a:ANSWER), (b:ANSWER)
                where b.x0 -(a.x0+a.x1)/2 > 0 and  b.x0 -(a.x0+a.x1)/2 < (case when a.x1-a.x0 > 20 Then a.x1-a.x0 else 20 end) //apoc.coll.max([a.x1-a.x0, 20]) //some rectangle not correctly bound the word --> heuristic
                     and abs(a.y0 - b.y0) < 18
                create (a)-[:NEXTWORD {y:abs(a.y0 - b.y0)}]->(b);
                ''')
                           
    tx.evaluate('''
                match (q:QUESTION), (a:ANSWER)
                where NOT (:ANSWER)-[:NEXTWORD]->(a)
                    and q.x1 < a.x0 and   a.x0 - q.x1 < 450
                    //and q.x1 < a.x0
                    and abs((q.y0+q.y1)/2 - (a.y0+a.y1)/2) < 10  //some rectangel is tall --> calculate average rectangle heigh (y)
                create (q)-[:RELATED {y:abs((q.y0+q.y1)/2 - (a.y0+a.y1)/2), x:(a.x0 - q.x1)}]->(a);
                ''')
    graph.commit(tx)
    
    #DELETE Redundant relations
    tx = graph.begin()
    tx.evaluate('''
                match (a:QUESTION)-[r1:NEXTWORD]->(b:QUESTION),
                    (a:QUESTION)-[r2:NEXTWORD]->(c:QUESTION)
                where r2.x > r1.x
                delete r2;
                ''')
                
    tx.evaluate('''
                match (a:ANSWER)-[r1:NEXTWORD]->(b:ANSWER),
                    (a:ANSWER)-[r2:NEXTWORD]->(c:ANSWER)
                where r2.y > r1.y
                delete r2;
                ''')
                
    tx.evaluate('''
                match (q1:QUESTION)-[r1:RELATED]->(a:ANSWER),
                     (q2:QUESTION)-[r2:RELATED]->(a:ANSWER)
                where r2.x > r1.x
                delete r2;
                ''')
    graph.commit(tx)
    
    ################# CONNECT ({word: 'Total'}) QUESTION --> Value ################
    tx = graph.begin()
    tx.evaluate('''
                match (q:QUESTION), (a:ANSWER) 
                where  ( (toLower(q.word) CONTAINS "total")  or (toLower(q.word) CONTAINS "summar") )
                    and q.x1 < a.x0
                    and abs((q.y0+q.y1)/2 - (a.y0+a.y1)/2) < 10  //some rectangel is tall --> calculate average rectangle heigh (y)
                create (q)-[:RELATED {y:abs((q.y0+q.y1)/2 - (a.y0+a.y1)/2), x:(a.x0 - q.x1)}]->(a);
                ''')
    graph.commit(tx)
    
    ################# CONNECT below ANS TO above QUESTION #########################
    #CASE:
    tx = graph.begin()    
    tx.evaluate('''
                match (q:QUESTION), (a:ANSWER)
                where not((q:QUESTION)-[]->(:QUESTION)) //end word in A question string
                    and NONE( x in ['PO NO','PONO'] where q.word contains x)
                    and not(()-[]->(a:ANSWER)) //start word in answer string
                    and abs(a.y0 - q.y1) < 35
                    and abs(a.x0 - q.x0) < $width/3.5
                create (q)-[:RELATED {y:abs(a.y0 - q.y1), x:abs(a.x1 - q.x0)}]->(a);
                ''', parameters = {'width': width})
    graph.commit(tx)
    
    #CASE: Summarize table (Quantity-Unit, G.W, N.W, Unit-Price, Amount)
    tx = graph.begin()
    tx.evaluate('''
                with ['Quantity', 'Unit', 'Net', 'Gross', 'AMOUNT','Goods'] as qung
                match (qstart:QUESTION), (qend:QUESTION), (a:ANSWER)
                where not(()-[]->(qstart:QUESTION)) //start word in A question string
                    and not ((qend:QUESTION)-[]->(:QUESTION)) //end word in A question string
                    and ((qstart:QUESTION)-[*0..]->(qend:QUESTION)) //a path from start 2 end (qstart->qend)
                    and not((:ANSWER)-[]->(a:ANSWER)) //start word in answer string
                    and a.y0 - qstart.y1 > 0 and a.y0 - qstart.y1 < 450
                    and abs(a.x0 - qstart.x0) < 45
                    and any (word in qung where qend.word contains word)
                    //and not exists ( OPTIONAL MATCH  (ansAboveA:ANSWER)
                    //                    where abs(ansAboveA.x0 - a.x0) < 5  and  0 < a.y0-ansAboveA.y0 )
                create (qend)-[:RELATED {y:abs(a.y0 - qend.y1), x:abs(a.x1 - qend.x0)}]->(a);
                ''', parameters = {'width': width})
    graph.commit(tx)
    
    tx = graph.begin()
    tx.evaluate('''
        match (q1:QUESTION)-[r1:RELATED]->(a:ANSWER),
              (q2:QUESTION)-[r2:RELATED]->(a:ANSWER)
        where r2.x > r1.x
        delete r2;
        ''')
    graph.commit(tx)
    
    ##################### FORCE some :OTHER to become :ANSWER #####################
    #####            Concate the right :OTHER to the left :ANSWER
    tx = graph.begin()
    tx.evaluate('''
                match p=(q:QUESTION) - [*1..] -> (a:ANSWER), (o:OTHER)
                where any(rel in relationships(p) WHERE type(rel) = "RELATED")
                    and not((:QUESTION)-[:NEXTWORD]->(q:QUESTION)) //get the longest path
                    and not((a:ANSWER)-[:NEXTWORD]->()) //get the longest path
                    and abs(a.y0 - o.y0) < 20 
                    and o.x0 - a.x1 > 0 //:OTHER must right of :ANSWER
                    and o.x0 - a.x1 <= 250
                REMOVE o:OTHER
                SET o:ANSWER
                create (a)-[:NEXTWORD {y:abs(a.y0 - o.y0)}]->(o);
                ''')
    graph.commit(tx)
    
    
    ################################   #GET RESULT  ############################
    query ='''
                match p=(q:QUESTION) - [*1..] -> (a:ANSWER)
                where any(rel in relationships(p) WHERE type(rel) = "RELATED")
                    and not((:QUESTION)-[:NEXTWORD]->(q:QUESTION)) //get the longest path
                    and not((a:ANSWER)-[:NEXTWORD]->()) //get the longest path
                return  trim(reduce(t="", str in [n in nodes(p) where 'QUESTION' in labels(n) | n.word] | t+str+" ")) as question,
                        trim(reduce(t="", str in [n in nodes(p) where 'ANSWER'   in labels(n) | n.word] | t+str+" ")) as answer
                order by a.y0 //order by from upper to lower text
                //order by length(p) DESC
                //limit 1
                ;   
           '''
    return graph.run(query).to_data_frame()

def invoice_info_extract_from_png(model, graph, path, tablePred, out_collection):
    fileFullnames = glob.glob(path + "\*.png")
    for file in fileFullnames:
        image = Image.open(file).convert("RGB")
        (width, height) = image.size
        # Transform image to force Tesseract recognize all text
        #(1198, 2022),Image.BICUBIC: miss Sailing on about
        #image = image.resize((965, 2022), Image.BICUBIC)  #(965, 2022)
        image = shear(image, 0.03)
        tableContainerSealNo = None
        strContainerSealNo = ""
        try:
            if "PACKING LIST" in file:
                tableContainerSealNo = tablePred.predict(image.copy())[-1]
                tableContainerSealNo.columns = tableContainerSealNo.iloc[0]
                tableContainerSealNo = tableContainerSealNo[1:]
                tableContainerSealNo[r'Container \ Seal No'] = tableContainerSealNo[r'Container'] + "\\" + tableContainerSealNo[r'Seal']
                strContainerSealNo = ' '.join(tableContainerSealNo[~tableContainerSealNo[r'Container \ Seal No'].isnull()][r'Container \ Seal No'])
        except:  #in case file doesn't have table of Container \ Seal No
            pass
        halfLeftImage = image.crop((0, 0, int(width/2), height))
        dfLeft = do_extract_from_png(halfLeftImage, file+'_left', model, graph)
        dfFull = do_extract_from_png(image, file+'_full', model, graph)
        df = pd.merge(dfLeft, dfFull, how='outer', left_on = 'question', right_on = 'question')
        df['answer'] = df.filter(like='answer').ffill(axis=1).iloc[:,-1]
        df = df[['question', 'answer']].drop_duplicates()
        df['question2header'] = ""
        df['sort'] = ""
        t = pd.DataFrame({'question':["Container \ Seal No"], 'question2header':["Container \ Seal No"], 'answer':[strContainerSealNo], 'sort':[14]})
        df = df.append(t)
        df.reset_index(drop=True, inplace=True)

        for index, row in df.iterrows():
            quest = row['question'].lower()
            if 'date' in quest and 'invoice' in quest:
                try: #dateutil.parser.
                    parse( row['answer'])
                    df.at[index,'question2header'] = "Invoice Date"
                    df.at[index,'sort'] = 2
                except:                    
                    df.at[index,'question2header'] = "Invoice No"
                    df.at[index,'sort'] = 1
            elif ('contract' in quest or 'invoice' in quest) and 'no' in quest:
                df.at[index,'question2header'] = "Invoice No"
                df.at[index,'sort'] = 1
            elif 'port' in quest and 'load' in quest:
                df.at[index,'question2header'] = "Arrival"
                df.at[index,'sort'] = 3
            elif 'final' in quest and 'dest' in quest:
                df.at[index,'question2header'] = "Departure"
                df.at[index,'sort'] = 4
            elif 'carrier' in quest:
                df.at[index,'question2header'] = "Carrier"
                df.at[index,'sort'] = 5
            elif 'sailing' in quest:
                try: #dateutil.parser
                    parse( row['answer'])
                    df.at[index,'question2header'] = "Sailing on About"
                    df.at[index,'sort'] = 6
                except:                    
                    df.at[index,'question2header'] = ""
            elif 'notify' in quest and 'party' in quest:
                df.at[index,'question2header'] = "Notify Party"
                df.at[index,'sort'] = 7
            elif 'shipper' in quest:
                df.at[index,'question2header'] = "Shipper"
                df.at[index,'sort'] = 8
            elif 'amount' in quest:
                df.at[index,'question2header'] = "Amount"
                df.at[index,'sort'] = 9
            elif 'unit' in quest:
                df.at[index,'question2header'] = "Unit"
                df.at[index,'sort'] = 10
            elif 'gross' in quest and ('weight' in quest or 'w' in quest):
                df.at[index,'question2header'] = "G.W"
                df.at[index,'sort'] = 11
            elif 'net' in quest and ('weight' in quest or 'w' in quest):
                df.at[index,'question2header'] = "N.W"
                df.at[index,'sort'] = 12
            elif 'description' in quest and 'goods' in quest:
                df.at[index,'question2header'] = "Description of Goods"
                df.at[index,'sort'] = 15
                
        out_collection[os.path.basename(file)] = df[['question', 'question2header', 'answer', 'sort']]
        

switcher = {
        1: "Invoice No",
        2: "Invoice Date",
        3: "Arrival",
        4: "Departure",
        5: "Carrier",
        6: "Sailing on About",
        7: "Notify Party",
        8: "Shipper",
        9: "Amount",
        10: "Unit",
        11: "G.W",
        12: "N.W",
        13: "Currency",
        14: "Container \ Seal No",
        15: "Description of Goods"
}
def findMissingData(arr):
    missing_elements = []
    for ele in range(1,len(switcher)+1):
        if ele not in arr:
            missing_elements.append(ele)
    return missing_elements

def numbers_to_header(argument):
    return switcher.get(argument, "nothing")

###################################### MAIN ###################################
from py2neo import Graph
from TableDetection import TablePredict
#uri = "neo4j+s://dff0ff04.databases.neo4j.io"
uri = "neo4j://localhost:7687"
user = "neo4j"
#password = "CJY__L21_jWwCzLqipfOM0WrAoPOH7LEomMCfN2GFN0"
password = "p@ss"
graph = Graph(uri, auth=(user, password))


import albumentations as album
from albumentations.pytorch.transforms import ToTensorV2

transforms = album.Compose([
    album.Resize(896, 896, always_apply=True),
    album.Normalize(),
    ToTensorV2()
])
tablePred = TablePredict(r'.\savedmodel\best_model.ckpt', transforms)

output_dictionary_png_df = {}
invoice_info_extract_from_png(model, graph, os.path.abspath(r".\InvoicesDirectory"), tablePred, output_dictionary_png_df)


################################### WRITE OUTPUT ##############################
from openpyxl import load_workbook, Workbook

book = Workbook() #load_workbook('./output.xlsx')
writer = pd.ExcelWriter(r'./InvoicesDirectory/output.xlsx', engine='openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
for df_name, df in output_dictionary_png_df.items():
    df.to_excel(writer, sheet_name=df_name)

writer.save()
writer.close()
print("Finish export output.xls")

################################# WRITE UPLOAD FILE ############################
output_dictionaryFinal = {}
for k, v in output_dictionary_png_df.items():
    file = k.split('.',1)[0]
    temp = v[['question2header', 'answer','sort']]
    temp.replace("", float("NaN"), inplace=True)
    temp.dropna(subset=['question2header'], inplace=True)
    temp.reset_index(drop=True, inplace=True)
    if file in output_dictionaryFinal:
        tempV = output_dictionaryFinal[file]
        finalPD = pd.merge(temp, tempV, how='outer', left_on = ['question2header','sort'], right_on = ['question2header','sort']).drop_duplicates()
        # lay data theo condtion
        finalPD['answer'] = ""
        currency = ""
        for index, row in finalPD.iterrows():
            quest = row['question2header']
            if quest in ['Departure', 'Arrival']:
                try:
                    finalPD.at[index,'answer'] = row['answer_x'] if len(row['answer_x']) > len(row['answer_y']) else row['answer_y']
                except:
                    finalPD.at[index,'answer'] = row['answer_x'] if pd.notna(row['answer_x']) else row['answer_y']
            elif quest in ['Carrier', 'Shipper']: #this condition is opposite with 'Invoice No'
                try:
                    finalPD.at[index,'answer'] = row['answer_y'] if len(row['answer_x']) > len(row['answer_y']) else row['answer_x']
                except:
                    finalPD.at[index,'answer'] = row['answer_x'] if pd.notna(row['answer_x']) else row['answer_y']
            elif 'Invoice No' in quest:
                ansx = (row['answer_y'][1:-1] if row['answer_y'][0] == 'I' else  row['answer_y']) if pd.isna(row['answer_x']) else (row['answer_x'][1:-1] if row['answer_x'][0] == 'I' else row['answer_x'])
                ansy = (row['answer_x'][1:-1] if row['answer_x'][0] == 'I' else  row['answer_x']) if pd.isna(row['answer_y']) else (row['answer_y'][1:-1] if row['answer_y'][0] == 'I' else row['answer_y'])
                finalPD.at[index,'answer'] = ansx if ansx == ansy else ansy
            elif quest in ['G.W', 'N.W', 'Notify Party', 'Sailing on About', 'Unit', 'Invoice Date', 'Container \ Seal No', 'Description of Goods']:
                finalPD.at[index,'answer'] = row['answer_y']  if pd.isna(row['answer_x']) else row['answer_x']
            elif quest in ['Amount']:
                ansx = "" if pd.isna(row['answer_x']) else row['answer_x']
                ansy = "" if pd.isna(row['answer_y']) else row['answer_y']
                amt = (ansx + " " + ansy).strip()
                finalPD.at[index,'answer'] = amt
                currency = Price.fromstring(amt).currency
        ####################### REFINE FINAL ##################################
        finalPD_addCurrency = finalPD[['question2header', 'answer', 'sort']].drop_duplicates()
        finalPD_addCurrency = finalPD_addCurrency.append(pd.DataFrame([["Currency", currency, 13]], columns=['question2header','answer','sort']))
        finalPD_addCurrency.reset_index(drop=True, inplace=True)
        
        temDF = finalPD_addCurrency.loc[finalPD_addCurrency['question2header'].isin(['Departure'])]      
        lengths = temDF["answer"].str.len()
        argmax = np.where(lengths == lengths.max())[0]        
        departDF = temDF.iloc[argmax]
        
        temDF = finalPD_addCurrency.loc[finalPD_addCurrency['question2header'].isin(['Shipper'])]
        lengths = temDF["answer"].str.len()
        argmin = np.where(lengths == lengths.min())[0]
        shipperDF = temDF.iloc[argmin]

        grossNetDF = finalPD_addCurrency.loc[finalPD_addCurrency['question2header'].isin(['G.W', 'N.W'])]
        # concatenate the string
        grossNetDF['answer'] = grossNetDF.groupby(['question2header'])['answer'].transform(lambda x : ', '.join(x))
        grossNetDF = grossNetDF.drop_duplicates()
        
        descriptionOfGoodsDF = finalPD_addCurrency.loc[finalPD_addCurrency['question2header'].isin(['Description of Goods'])]
        descriptionOfGoodsDF = descriptionOfGoodsDF[0:-1] #Delete the thired row, only accept the first 2 rows
        descriptionOfGoodsDF['answer'] = descriptionOfGoodsDF.groupby(['question2header'])['answer'].transform(lambda x : ', '.join(x))
        descriptionOfGoodsDF = descriptionOfGoodsDF.drop_duplicates()
        
        unitDF = finalPD_addCurrency.loc[finalPD_addCurrency['question2header'].isin(['Unit'])]
        unitDF = unitDF[unitDF['answer'].str.contains('\d')]

        restDF =  finalPD_addCurrency.loc[~finalPD_addCurrency['question2header'].isin(['Departure','G.W','N.W','Unit','Shipper','Description of Goods'])]
        refinedFinalDF = departDF.append(grossNetDF).append(descriptionOfGoodsDF).append(restDF).append(shipperDF).append(unitDF)
        missData = findMissingData(refinedFinalDF['sort'].tolist())
        for miss in missData:
            refinedFinalDF = refinedFinalDF.append(pd.DataFrame([[numbers_to_header(miss), "", miss]], columns=['question2header','answer','sort']))
        refinedFinalDF = refinedFinalDF.append(pd.DataFrame([["File", file, 0]], columns=['question2header','answer','sort']))
        refinedFinalDF = refinedFinalDF.sort_values(by=['sort'])
        refinedFinalDF.reset_index(drop=True, inplace=True)
        output_dictionaryFinal[file] = refinedFinalDF
    else:
        output_dictionaryFinal[file] = temp
        

exportDF = None
for df_name, df in output_dictionaryFinal.items():
    if exportDF is None:
        t = df[['question2header', 'answer']]
        t.set_index('question2header', inplace=True)
        exportDF = t.T
    else:
        t = df[['question2header', 'answer']]
        t.set_index('question2header', inplace=True)
        exportDF = exportDF.append(t.T)

exportDF.to_excel(r'.\InvoicesDirectory\outputFinal.xls')

# book = Workbook() #load_workbook('./output.xlsx')
# writer = pd.ExcelWriter(r'./InvoicesDirectory/output.xlsx', engine='openpyxl')
# writer.book = book
# writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
# for df_name, df in output_dictionary_png_df.items():
#     df.to_excel(writer, sheet_name='Sheet 1')

# writer.save()
# writer.close()        



print("Finish export outputFinal.xls")

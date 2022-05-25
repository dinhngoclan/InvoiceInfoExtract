import os
#os.system('pip install git+https://github.com/huggingface/transformers.git --upgrade')
#os.system('pip install pyyaml==5.1')
# workaround: install old version of pytorch since detectron2 hasn't released packages for pytorch 1.9 (issue: https://github.com/facebookresearch/detectron2/issues/3158)
#os.system('pip install torch==1.8.0+cu101 torchvision==0.9.0+cu101 -f https://download.pytorch.org/whl/torch_stable.html')

# install detectron2 that matches pytorch 1.8
# See https://detectron2.readthedocs.io/tutorials/install.html for instructions
#os.system('pip install -q detectron2 -f https://dl.fbaipublicfiles.com/detectron2/wheels/cu101/torch1.8/index.html')

## install PyTesseract
#os.system('pip install -q pytesseract')



import numpy as np
import pandas as pd
from transformers import LayoutLMv2FeatureExtractor, LayoutLMv2TokenizerFast, LayoutLMv2Tokenizer, LayoutLMv2ForTokenClassification
from PIL import Image, ImageDraw, ImageFont, ImageOps
import datefinder

feature_extractor = LayoutLMv2FeatureExtractor.from_pretrained(r".\savedmodel")#("microsoft/layoutlmv2-base-uncased")
tokenizer = LayoutLMv2TokenizerFast.from_pretrained(r".\savedmodel")#("microsoft/layoutlmv2-base-uncased")
model = LayoutLMv2ForTokenClassification.from_pretrained(r".\savedmodel")

# define id2label, label2color
#labels = dataset.features['ner_tags'].feature.names
labels = ['O', 'B-HEADER', 'I-HEADER', 'B-QUESTION', 'I-QUESTION', 'B-ANSWER', 'I-ANSWER']
id2label = {v: k for v, k in enumerate(labels)}
label2color = {'question':'blue', 'answer':'green', 'header':'orange', 'other':'violet'}

def unnormalize_box(bbox, width, height):
     return [
         width * (bbox[0] / 1000),
         height * (bbox[1] / 1000),
         width * (bbox[2] / 1000),
         height * (bbox[3] / 1000),
     ]

def iob_to_label(label):
    label = label[2:]
    if not label:
      return 'other'
    return label

def shear(img, shear):
    shear = img.transform(img.size, Image.AFFINE, (1, shear, 0, 0, 1, 0))
    return shear

def process_image(model, image):
    width, height = image.size
    
    # get words, boxes
    encoding_feature_extractor = feature_extractor(image, return_tensors="pt")
    words, boxes = encoding_feature_extractor.words, encoding_feature_extractor.boxes

    # encode
    encoding = tokenizer(words, boxes=boxes, return_offsets_mapping=True, return_tensors="pt")
    offset_mapping = encoding.pop('offset_mapping')
    encoding["image"] = encoding_feature_extractor.pixel_values

    # forward pass
    outputs = model(**encoding)

    # get predictions
    predictions = outputs.logits.argmax(-1).squeeze().tolist()
    token_boxes = encoding.bbox.squeeze().tolist()

    # only keep non-subword predictions
    is_subword = np.array(offset_mapping.squeeze().tolist())[:,0] != 0
    true_predictions = [id2label[pred] for idx, pred in enumerate(predictions) if not is_subword[idx]]
    true_boxes = [unnormalize_box(box, width, height) for idx, box in enumerate(token_boxes) if not is_subword[idx]]

  
    origBoxes = []
    for box in boxes[0] : 
        origBoxes.append(unnormalize_box(box, width, height))
    
    pdWordsBoxes = pd.concat([pd.DataFrame({'words': words[0]}), pd.DataFrame({'origBoxes': origBoxes})], axis=1, join='inner')
    pdWordsBoxes['origBoxes_STR'] = pdWordsBoxes['origBoxes'].astype(str)
    pdPred = pd.DataFrame(zip(true_boxes, true_predictions),columns = ['true_boxes', 'true_predictions'])
    pdPred['true_boxes_STR'] = pdPred['true_boxes'].astype(str)
    finalPD = pd.merge(pdWordsBoxes, pdPred, how='left', left_on = 'origBoxes_STR', right_on = 'true_boxes_STR')
    finalPD = finalPD.dropna(subset = ['true_predictions'])
    finalPD = finalPD[~finalPD.words.isin(["&", ",", "."]) ]
    
    for index, row in finalPD.iterrows():
        try:
            is_date = False
            for match in datefinder.find_dates( row['words']):
                is_date = True
            if is_date and 'ANSWER' not in row['true_predictions']:
                finalPD.at[index, 'true_predictions'] = "I-ANSWER"
        except:
            pass
    
    
    # draw predictions over the image
    draw = ImageDraw.Draw(image)
    font = ImageFont.load_default()
    for prediction, box in zip(finalPD['true_predictions'], finalPD['true_boxes']):
        #print(prediction)
        predicted_label = iob_to_label(prediction).lower()
        draw.rectangle(box, outline=label2color[predicted_label])
        draw.text((box[0]+10, box[1]-10), text=predicted_label, fill=label2color[predicted_label], font=font)
  
    return image, finalPD


import glob
def do_extract_from_png(image, file, model, graph):
    width, height = image.size
    #===================== Use ML to classify word ========================
    image, table = process_image(model, image)
    image.save(file+'.jpg')
    #================ Use Graph to math Question-Answer ===================
    tx = graph.begin()
    tx.evaluate('''  MATCH (n) DETACH DELETE n  ''')
    graph.commit(tx)
    #CREATE NODES
    tx = graph.begin()
    for index, row in table.iterrows():
        try:
            if row['true_predictions'] == 'O':
                tx.evaluate('''
                   MERGE (o:OTHER  {word:$label2, x0:$x0, y0:$y0, x1:$x1, y1:$y1})
                   ''', parameters = {'label2': row['words'], 'x0':row['origBoxes'][0], 'y0':row['origBoxes'][1]
                                                           , 'x1':row['origBoxes'][2], 'y1':row['origBoxes'][3]})
            elif 'QUESTION' in row['true_predictions']:
                tx.evaluate('''
                   MERGE (q:QUESTION {word:$label2, x0:$x0, y0:$y0, x1:$x1, y1:$y1})
                   ''', parameters = {'label2': row['words'], 'x0':row['origBoxes'][0], 'y0':row['origBoxes'][1]
                                                           , 'x1':row['origBoxes'][2], 'y1':row['origBoxes'][3]})
            elif 'ANSWER' in row['true_predictions']:
                tx.evaluate('''
                   MERGE (a:ANSWER {word:$label2, x0:$x0, y0:$y0, x1:$x1, y1:$y1})
                   ''', parameters = {'label2': row['words'], 'x0':row['origBoxes'][0], 'y0':row['origBoxes'][1]
                                                           , 'x1':row['origBoxes'][2], 'y1':row['origBoxes'][3]})
        except:
            print(row['true_predictions'])
        
    graph.commit(tx)
    
    #CREATE RELATIONSHIPs
    tx = graph.begin()
    tx.evaluate('''
                match (a:QUESTION), (b:QUESTION)
                where a.x0 - b.x0 < 0 and  a.x1 - b.x0 > -20
                     and abs(a.y0 - b.y0) < 10 
                create (a)-[:NEXTWORD {x:abs(a.x1-b.x0)}]->(b);
                ''')
                
    tx.evaluate('''
                match (a:ANSWER), (b:ANSWER)
                where b.x0 -(a.x0+a.x1)/2 > 0 and  b.x0 -(a.x0+a.x1)/2 < (case when a.x1-a.x0 > 20 Then a.x1-a.x0 else 20 end) //apoc.coll.max([a.x1-a.x0, 20]) //some rectangle not correctly bound the word --> heuristic
                     and abs(a.y0 - b.y0) < 18
                create (a)-[:NEXTWORD {y:abs(a.y0 - b.y0)}]->(b);
                ''')
                           
    tx.evaluate('''
                match (q:QUESTION), (a:ANSWER)
                where NOT (:ANSWER)-[:NEXTWORD]->(a)
                    and q.x1 < a.x0 and   a.x0 - q.x1 < 450
                    //and q.x1 < a.x0
                    and abs((q.y0+q.y1)/2 - (a.y0+a.y1)/2) < 10  //some rectangel is tall --> calculate average rectangle heigh (y)
                create (q)-[:RELATED {y:abs((q.y0+q.y1)/2 - (a.y0+a.y1)/2), x:(a.x0 - q.x1)}]->(a);
                ''')
    graph.commit(tx)
    
    #DELETE Redundant relations
    tx = graph.begin()
    tx.evaluate('''
                match (a:QUESTION)-[r1:NEXTWORD]->(b:QUESTION),
                    (a:QUESTION)-[r2:NEXTWORD]->(c:QUESTION)
                where r2.x > r1.x
                delete r2;
                ''')
                
    tx.evaluate('''
                match (a:ANSWER)-[r1:NEXTWORD]->(b:ANSWER),
                    (a:ANSWER)-[r2:NEXTWORD]->(c:ANSWER)
                where r2.y > r1.y
                delete r2;
                ''')
                
    tx.evaluate('''
                match (q1:QUESTION)-[r1:RELATED]->(a:ANSWER),
                     (q2:QUESTION)-[r2:RELATED]->(a:ANSWER)
                where r2.x > r1.x
                delete r2;
                ''')
    graph.commit(tx)
    
    ################# CONNECT ({word: 'Total'}) QUESTION --> Value ################
    tx = graph.begin()
    tx.evaluate('''
                match (q:QUESTION), (a:ANSWER)
                where  ( (toLower(q.word) CONTAINS "total")  or (toLower(q.word) CONTAINS "summar") )
                    and q.x1 < a.x0
                    and abs((q.y0+q.y1)/2 - (a.y0+a.y1)/2) < 10  //some rectangel is tall --> calculate average rectangle heigh (y)
                create (q)-[:RELATED {y:abs((q.y0+q.y1)/2 - (a.y0+a.y1)/2), x:(a.x0 - q.x1)}]->(a);
                ''')
    graph.commit(tx)
    
    ################# CONNECT below ANS TO above QUESTION #########################
    tx = graph.begin()
    tx.evaluate('''
                match (q:QUESTION), (a:ANSWER)
                where not((q:QUESTION)-[]->(:QUESTION)) //end word in A question string
                    and not(()-[]->(a:ANSWER)) //start word in answer string
                    and abs(a.y0 - q.y1) < 35
                    and abs(a.x0 - q.x0) < $width/3.5
                create (q)-[:RELATED {y:abs(a.y0 - q.y1), x:abs(a.x1 - q.x0)}]->(a);
                ''', parameters = {'width': width})
    graph.commit(tx)
    
    tx = graph.begin()
    tx.evaluate('''
        match (q1:QUESTION)-[r1:RELATED]->(a:ANSWER),
              (q2:QUESTION)-[r2:RELATED]->(a:ANSWER)
        where r2.x > r1.x
        delete r2;
        ''')
    graph.commit(tx)
    
    ##################### FORCE some :OTHER to become :ANSWER #####################
    #####            Concate the right :OTHER to the left :ANSWER
    tx = graph.begin()
    tx.evaluate('''
                match p=(q:QUESTION) - [*1..] -> (a:ANSWER), (o:OTHER)
                where any(rel in relationships(p) WHERE type(rel) = "RELATED")
                    and not((:QUESTION)-[:NEXTWORD]->(q:QUESTION)) //get the longest path
                    and not((a:ANSWER)-[:NEXTWORD]->()) //get the longest path
                    and abs(a.y0 - o.y0) < 20 
                    and o.x0 - a.x1 > 0 //:OTHER must right of :ANSWER
                    and o.x0 - a.x1 <= 250
                REMOVE o:OTHER
                SET o:ANSWER
                create (a)-[:NEXTWORD {y:abs(a.y0 - o.y0)}]->(o);
                ''')
    graph.commit(tx)
    
    
    ################################   #GET RESULT  ############################
    query ='''
                match p=(q:QUESTION) - [*1..] -> (a:ANSWER)
                where any(rel in relationships(p) WHERE type(rel) = "RELATED")
                    and not((:QUESTION)-[:NEXTWORD]->(q:QUESTION)) //get the longest path
                    and not((a:ANSWER)-[:NEXTWORD]->()) //get the longest path
                return  trim(reduce(t="", str in [n in nodes(p) where 'QUESTION' in labels(n) | n.word] | t+str+" ")) as question,
                        trim(reduce(t="", str in [n in nodes(p) where 'ANSWER'   in labels(n) | n.word] | t+str+" ")) as answer;   
           '''
    return graph.run(query).to_data_frame() 
    
def invoice_info_extract_from_png(model, graph, path, out_collection):
    fileFullnames = glob.glob(path + "\*.png")
    for file in fileFullnames:
        image = Image.open(file).convert("RGB")
        (width, height) = image.size
        # Transform image to force Tesseract recognize all text
        #(1198, 2022),Image.BICUBIC: miss Sailing on about
        #image = image.resize((965, 2022), Image.BICUBIC)  #(965, 2022)
        #image = shear(image, 0.03)
        halfLeftImage = image.crop((0, 0, int(width/2), height))
        #image = ImageOps.grayscale(image)
        dfLeft = do_extract_from_png(halfLeftImage, file+'_left', model, graph)
        dfFull = do_extract_from_png(image, file+'_full', model, graph) 
        df = pd.merge(dfLeft, dfFull, how='outer', left_on = 'question', right_on = 'question')
        df['answer'] = df.filter(like='answer').ffill(axis=1).iloc[:,-1]	
        out_collection[os.path.basename(file)] = df[['question', 'answer']].drop_duplicates()

###################################### MAIN ###################################
from py2neo import Graph
#uri = "neo4j+s://dff0ff04.databases.neo4j.io"
uri = "neo4j://localhost:7687"
user = "neo4j"
#password = "CJY__L21_jWwCzLqipfOM0WrAoPOH7LEomMCfN2GFN0"
password = "p@SS123$%^"
graph = Graph(uri, auth=(user, password))

output_dictionary = {} 
invoice_info_extract_from_png(model, graph, os.path.abspath(r".\InvoicesDirectory"), output_dictionary)


################################### WRITE OUTPUT ##############################
from openpyxl import load_workbook, Workbook

book = Workbook() #load_workbook('./output.xlsx')
writer = pd.ExcelWriter(r'./InvoicesDirectory/output.xlsx', engine='openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

for df_name, df in output_dictionary.items():
    df.to_excel(writer, sheet_name=df_name)

writer.save()
writer.close()
   

# -*- coding: utf-8 -*-
"""
Created on Wed Jul  6 08:52:26 2022

@author: samsung
"""

import numpy as np
import pandas as pd
import re
from transformers import LayoutLMv2FeatureExtractor, LayoutLMv2TokenizerFast, LayoutLMv2Tokenizer, LayoutLMv2ForTokenClassification
from ApplyLayoutLMv2_Graph_TableNet_Util import *
from py2neo import Graph
from TableDetection import TablePredict
import albumentations as album
from albumentations.pytorch.transforms import ToTensorV2
import datetime

switcherEinvoice = {
        1: "Invoice No",
        2: "Invoice Date",
        3: "CompanyName",
        4: "Address",
        5: "GoodsDescription",
        6: "TotalAmount",
        7: "VatAmount",
        8: "GrandTotal"
}

def question2header_func(df):
    for index, row in df.iterrows():
        quest = row['question'].lower()
        if 'ngày' in quest or 'ngay' in quest or 'date' in quest:
            try: #dateutil.parser.
                #parse( row['answer']) #currently cannot parse "23 tháng (month) 06 năm (year) 2022"
                df.at[index,'question2header'] = "Invoice Date"
                dates = [str(s) for s in row['answer'].split() if s.isdigit()]
                strDate = '/'.join(dates)
                try:
                    df.at[index,'answer'] = datetime.datetime.strptime(strDate,'%d/%m/%Y').strftime('%d-%b-%Y')
                except: 
                    df.at[index,'answer'] = strDate
                    
                df.at[index,'sort'] = 2
            except:                     
                df.at[index,'question2header'] = "Invoice No"
                df.at[index,'sort'] = 1       
        elif 'số (no' in quest or 'sé' in quest or 'so' in quest or 'sô' in quest or 'số /wo' in quest or 'só' in quest or 'số (wo' in quest:
            df.at[index,'question2header'] = "Invoice No"
            df.at[index,'sort'] = 1
        elif 'tên đơn vị' in quest or 'đơn vị' in quest  or 'company name' in quest:
            df.at[index,'question2header'] = "CompanyName"
            df.at[index,'sort'] = 3
        elif 'địa chỉ' in quest and 'address' in quest:
            df.at[index,'question2header'] = "Address"
            df.at[index,'sort'] = 4

    return df      

###################################### MAIN ###################################

#======================Convert .pdf to .png ========================
from pdf2image import convert_from_path
filenames = glob.glob('.\InvoicesDirectory' + "\*.pdf")
for file in filenames:
    images = convert_from_path(file,poppler_path=r'.\poppler-22.01.0\Library\bin')
    (width, height) = images[0].size
    halfTopImage = images[0].crop((0, 0, width, int(height/2)))
    halfTopImage.save(file+'.png', "PNG")
    img = remove_black_dot3(file+'.png')
    
    
def extQuestionAnswerGraphFunc(graph, file=None, tablePD=None, considerWidth=None):
    #DELETE Redundant relations
    tx = graph.begin()                            
    tx.evaluate('''
                match (q1:QUESTION)-[r1:RELATED]->(a:ANSWER),
                     (q2:QUESTION)-[r2:RELATED]->(a:ANSWER)
                where r2.x > r1.x 
                    and 0 < q2.y0-10 <= q1.y0 <= q2.y0+10
                delete r2;
                ''')
    graph.commit(tx)
     
    ################# CONNECT below ANS TO above QUESTION #########################
    #CASE:
    try:
        filter1 = table['words'].str.contains('descript|Descript|DESCRIPT')
        description_y0 = table.where(filter1).dropna().reset_index(drop=True)['origBoxes'].iloc[0][1]
    except:
        description_y0 = 950
    tx = graph.begin()
    tx.evaluate('''
                with ['quantity', 'unit', 'net', 'gross', 'amount','total','po no','pono','no'] as not_map_now //Unit here is UnitPrice
                match (q:QUESTION), (a:ANSWER)
                where not((q:QUESTION)-[]->(:QUESTION)) //end word in A question string
                    and not any (word in not_map_now where toLower(q.word) contains word)
                    //and NONE( x in ['PO NO','PONO','NO'] where q.word contains x) //some time 'PO' 'NO' are on 2 different nodes
                    and not((:ANSWER)-[]->(a:ANSWER)) //start word in answer string
                    and abs(a.y0 - q.y1) < 35
                    and abs(a.x0 - q.x0) < $width/3.5
                    and q.y0 < $description_y0 // 950 date modify: 20220725
                create (q)-[:RELATED {y:abs(a.y0 - q.y1), x:abs(a.x1 - q.x0)}]->(a);
                ''', parameters = {'width': considerWidth, 'description_y0':description_y0})
    graph.commit(tx)
    
    #CASE: Summarize table (Description of Goods)
    focusHigh = 550
    if file:
        if "INVOICE" in file or "bottom" in file:     
            focusHigh = 650
   
    
    tx = graph.begin()
    tx.evaluate('''
        with ['quantity', 'unit', 'net', 'gross', 'amount','goods', 'total'] as qung //Unit here is UnitPrice, [20220802]
        match (q1:QUESTION)-[r1:RELATED]->(a:ANSWER),
              (q2:QUESTION)-[r2:RELATED]->(a:ANSWER)
        where r2.x > r1.x
            and not any (word in qung where toLower(q2.word) contains word)
            and 0 < q2.y0-10 <= q1.y0 <= q2.y0+10
        delete r2;
        ''')
    graph.commit(tx)
    
    tx = graph.begin()
    tx.evaluate('''
        match (a1:ANSWER)-[r1:NEXTWORD]->(a11:ANSWER),
              (a1:ANSWER)-[r2:NEXTWORD]->(a12:ANSWER)
        where abs(a11.y0 - a12.y0) < 3 and abs(a11.y1 - a12.y1) < 3 
            and a11.x1 < a12.x0
        delete r2;
        ''')
    graph.commit(tx)
    
    
    ##################### FORCE some :OTHER to become :ANSWER #####################
    #####            Concate the right :OTHER to the left :ANSWER
    tx = graph.begin()
    tx.evaluate('''
                match p=(q:QUESTION) - [*1..] -> (a:ANSWER), (o:OTHER)
                where any(rel in relationships(p) WHERE type(rel) = "RELATED")
                    and not((:QUESTION)-[:NEXTWORD]->(q:QUESTION)) //get the longest path
                    and not((a:ANSWER)-[:NEXTWORD]->()) //get the longest path
                    and abs(a.y0 - o.y0) < 20 
                    and o.x0 - a.x1 > 0 //:OTHER must right of :ANSWER
                    and o.x0 - a.x1 <= 250
                REMOVE o:OTHER
                SET o:ANSWER
                create (a)-[:NEXTWORD {y:abs(a.y0 - o.y0)}]->(o);
                ''')
    graph.commit(tx)
    
    tx = graph.begin()             
    tx.evaluate('''
                match (q1:QUESTION)-[r1:RELATED]->(a:ANSWER),
                     (q2:QUESTION)-[r2:RELATED]->(a:ANSWER)
                where r2.y > r1.y+4.2
                delete r2;
                ''')
    graph.commit(tx)

feature_extractor = LayoutLMv2FeatureExtractor.from_pretrained(r".\savedmodel", ocr_lang='vie')#("microsoft/layoutlmv2-base-uncased")
tokenizer = LayoutLMv2TokenizerFast.from_pretrained(r".\savedmodel")#("microsoft/layoutlmv2-base-uncased")
model = LayoutLMv2ForTokenClassification.from_pretrained(r".\savedmodel")

#uri = "neo4j+s://dff0ff04.databases.neo4j.io"
uri = "neo4j://localhost:7687"
user = "neo4j"
password = "CJY__L21_jWwCzLqipfOM0WrAoPOH7LEomMCfN2GFN0"

graph = Graph(uri, auth=(user, password))


# transforms = album.Compose([
#     album.Resize(896, 896, always_apply=True),
#     album.Normalize(),
#     ToTensorV2()
# ])
#tablePred = TablePredict(r'.\savedmodel\best_model.ckpt', transforms, threshold=0.3, lang='vie')

output_dictionary_png_df = {}
lstHasToBeQuestion=["Ngày","Ngay", "Date", "Dafe", "Ma", "Mã", "thuê", "thuế", "Payment", "Method", "Currency", "Company", "Name"]
lstHasToBeAnswer=["(month)","tháng","(year)","năm"]
invoice_info_extract_from_png(feature_extractor, tokenizer, model, graph, os.path.abspath(r".\InvoicesDirectory"), question2header_func, 
                              switcher = switcherEinvoice, tablePred=None, out_collection=output_dictionary_png_df, 
                              cut_image=None, image_shear=0.02,#0.035,
                              lstHasToBeQuestion=lstHasToBeQuestion,
                              lstHasToBeAnswer=lstHasToBeAnswer,
                              extQuestionAnswerGraphFunc=extQuestionAnswerGraphFunc)

import tabula
fileFullnames = glob.glob(r".\InvoicesDirectory" + "\*.pdf")
for file in fileFullnames:
    dfs = tabula.read_pdf(file, pages='all')
    s =  dfs[0].iloc[:,0]
    try:
        goodsDescription = s.loc[s.str.startswith(('1','2','3','4','5','6','7','8','9','10'),na=False)].str.cat(sep='\n') #.iloc[0]
    except: 
        goodsDescription = s.iloc[2]
    if not bool(re.match(r'^[\d]+[ \w]+', goodsDescription)) and s[2] and s[3] and s[4]:
        goodsDescription = '1 '+s[2]+' '+s[4] 
    df = dfs[0].dropna(subset=[r"Đơn vị tính Số lượng"])
    totalAmt = df[df[r"Đơn vị tính Số lượng"].str.contains("Total")][r"Thành tiền"].iloc[0]
    vatAmt = df[df[r"Đơn vị tính Số lượng"].str.contains("VAT")][r"Thành tiền"].iloc[0]
    grandTotal = df[df[r"Đơn vị tính Số lượng"].str.contains("Grand")][r"Thành tiền"].iloc[0]
    
    
    tempV = output_dictionary_png_df[os.path.basename(file)+'.png']
    tempV = tempV.append(pd.DataFrame([["GoodsDescription", goodsDescription, 5]], columns=['question2header','answer','sort']))
    tempV = tempV.append(pd.DataFrame([["TotalAmount", float(totalAmt.replace(',','')), 6]], columns=['question2header','answer','sort']))
    tempV = tempV.append(pd.DataFrame([["VatAmount", float(vatAmt.replace(',','')), 7]], columns=['question2header','answer','sort']))
    tempV = tempV.append(pd.DataFrame([["GrandTotal", float(grandTotal.replace(',','')), 8]], columns=['question2header','answer','sort']))
    output_dictionary_png_df[os.path.basename(file)+'.png'] = tempV
################################### WRITE OUTPUT ##############################
from openpyxl import load_workbook, Workbook

book = Workbook() #load_workbook('./output.xlsx')
writer = pd.ExcelWriter(r'./InvoicesDirectory/output.xlsx', engine='openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
for df_name, df in output_dictionary_png_df.items():
    df.to_excel(writer, sheet_name=df_name)

writer.save()
writer.close()
print("Finish export output.xls")

################################# WRITE UPLOAD FILE ############################
output_dictionaryFinal = {}
for k, v in output_dictionary_png_df.items():
    file = k.split('.',1)[0]
    temp = v[['question2header', 'answer','sort']]
    temp.replace("", float("NaN"), inplace=True)
    temp.dropna(subset=['question2header'], inplace=True)
    temp.reset_index(drop=True, inplace=True)
    missData = findMissingData(temp['sort'].tolist(), switcher=switcherEinvoice)
    for miss in missData:
        temp = temp.append(pd.DataFrame([[numbers_to_header(miss, switcher=switcherEinvoice), "", miss]], columns=['question2header','answer','sort']))
    temp = temp.sort_values(by=['sort'])
    temp.reset_index(drop=True, inplace=True)
    
    temDF2 = temp.loc[temp['question2header'].isin(['Invoice Date'])]
    lengths = temDF2["answer"].str.len()
    argmax = np.where(lengths == lengths.max())[0]        
    invoiceDateDF = temDF2.iloc[argmax]
    
    restDF =  temp.loc[~temp['question2header'].isin(['Invoice Date'])]
    refinedFinalDF =  invoiceDateDF.append(restDF)
    refinedFinalDF = refinedFinalDF.sort_values(by=['sort'])
    refinedFinalDF.reset_index(drop=True, inplace=True)
    output_dictionaryFinal[file] = refinedFinalDF
        

exportDF = None
for df_name, df in output_dictionaryFinal.items():
    if exportDF is None:
        t = df[['question2header', 'answer']]
        t.set_index('question2header', inplace=True)
        exportDF = t.T
    else:
        t = df[['question2header', 'answer']]
        t.set_index('question2header', inplace=True)
        exportDF = exportDF.append(t.T)

exportDF.to_excel(r'.\InvoicesDirectory\outputEinvoiceFinal.xls')

# book = Workbook() #load_workbook('./output.xlsx')
# writer = pd.ExcelWriter(r'./InvoicesDirectory/output.xlsx', engine='openpyxl')
# writer.book = book
# writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
# for df_name, df in output_dictionary_png_df.items():
#     df.to_excel(writer, sheet_name='Sheet 1')

# writer.save()
# writer.close()        



print("Finish export outputEinvoiceFinal.xls")
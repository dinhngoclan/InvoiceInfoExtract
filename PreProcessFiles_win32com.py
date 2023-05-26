# -*- coding: utf-8 -*-
"""
Created on Thu Apr 28 17:05:24 2022

@author: ngoclan
"""

#=======================Convert .xls to .xlsx========================
import os, glob
import win32com.client as win32
from win32com import client
client.win32com_delete_gen_path = 1

import time

#import warnings
#warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
# Excel object must be global and Closing at the end of programe, instead of redefine and close in each function  
# --> to avoid error 'client disconected' 
excel = win32.gencache.EnsureDispatch('Excel.Application')

def convert_xls2xlsx(path):
    #excel = win32.gencache.EnsureDispatch('Excel.Application')
    #excel = client.Dispatch("Excel.Application")
    filenames = glob.glob(path + "\*.xls")
    filenames = [f for f in filenames if "~" not in f] # remove temporary files
    for file in filenames:
        #os.path.abspath(r".\TCDKO-2204-08 120422 (Kumho-2309)-01 NYLON-BUSAN-XKD.XLS")
        print ("convert_xls2xlsx: ",file)
        wb = excel.Workbooks.Open(file)
        #time.sleep(0.5)
        
        wb.SaveAs(file+'x', FileFormat = 51)    #FileFormat = 51 is for .xlsx extension        
        wb.Close()                               #FileFormat = 56 is for .xls extension
        wb = None
    #excel.Application.Quit()

convert_xls2xlsx(os.path.abspath(r".\InvoicesDirectory"))

#=======================Hide GridLine in .xlsx ========================
##OK OK
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles.alignment import Alignment
# Load workbook
import pyexcel
#pyexcel.save_book_as(file_name='.\TCDKO-2204-08 120422 (Kumho-2309)-01 NYLON-BUSAN-XKD.XLS', dest_file_name='.\TCDKO-2204-08 120422 (Kumho-2309)-01 NYLON-BUSAN-XKD.XLSX')
def hide_gridline_in_xlsx(path):
    filenames = glob.glob(path + "\*.xlsx")
    filenames = [f for f in filenames if "~" not in f] # remove temporary files
    for file in filenames:
        print('hide_gridline_in_xlsx: ', file)
        wb = openpyxl.load_workbook(file)
        try:
            sheetDoc = wb['DOC']
            packingList = wb.copy_worksheet(sheetDoc)
            invoice = wb.copy_worksheet(sheetDoc)
            
            packingList.title = 'PACKING LIST'
            #packingList.row_dimensions.group(55,250, hidden=True) #.delete_rows(55,2500)
            for r in range(1,packingList.max_row+1):
                for c in range(1,packingList.max_column+1):
                    s = packingList.cell(r,c).value
                    try:
                        if s and ("PACKING" in s and "LIST" in s):
                            packingList.row_dimensions.group(1,r-1, hidden=True)
                            packingList.row_dimensions.group(r+57,packingList.max_row, hidden=True)
                            break
                    except Exception as e:
                        pass
            
            
            invoice.title = 'INVOICE'
            for r in range(1,invoice.max_row+1):
                for c in range(1,invoice.max_column+1):
                    s = invoice.cell(r,c).value
                    try:
                        if s and ("COMMERCIAL" in s and "INVOICE" in s):  #"COMMERCIAL INVOICE" in s: 
                            #invoice.delete_rows(1,r-1)
                            invoice.row_dimensions.group(1,r-1, hidden=True)
                            invoice.row_dimensions.group(r+57,invoice.max_row, hidden=True)
                            # Save workbook
                            wb.save(file)
                            break
                    except:
                        pass            
        except:
            pass
        # Loop through all cells in all worksheets
        for sheet in wb.worksheets:
            sheet.sheet_view.showGridLines = False
            sheet.sheet_view.view='normal'
            sheet.paper_size_code = 1
            #sheet.title = sheet.title.strip()
            #print(sheet.title)
            #sheet.alignment = Alignment(horizontal="center")
            #remove "///////////////////////////"
            for r in range(1,sheet.max_row+1):
                for c in range(1,sheet.max_column+1):
                    if 'H_C' in sheet.cell(r,c).font.name:
                        sheet.cell(r,c).font =  openpyxl.styles.Font(name='Arial') #= sheet.cell(r,c).style.copy(font= 'Arial')
                    s = sheet.cell(r,c).value
                    try:
                        if s != None and "/////////////////////////////////////////////////////////////" in s: 
                            sheet.cell(r,c).value = ''
                        elif  s != None and "AMOUNT" in s and sheet.column_dimensions[get_column_letter(c)].width < 22:
                            sheet.column_dimensions[get_column_letter(c)].width = 22
                    except Exception as e:
                        pass
            # Save workbook
            wb.save(file)
            
        wb.close()
 
#####hide_gridline_in_xlsx(os.path.abspath(r".\InvoicesDirectory"))


###############################################################################
def parse_sheet_doc_usingWin32Com(path):
    #excel = client.Dispatch("Excel.Application") 
    filenames = glob.glob(path + "\*.xlsx")
    filenames = [f for f in filenames if "~" not in f] # remove temporary files
    for file in filenames:
        print('parse_sheet_doc_usingWin32Com: ', file)
        #wb = excel.Workbooks.Open(file)
        try:
            wb = excel.Workbooks.Open(file)
            sheetDoc =  wb.Sheets['DOC']
            sheetDoc.Copy(None, wb.Sheets(wb.Sheets.Count))
            #packingList = wb.Sheets(wb.Sheets.Count)
            packingList = wb.Sheets['DOC (2)']
            packingList.Name = 'PACKING LIST'
            
            sheetDoc.Copy(None, wb.Sheets(wb.Sheets.Count))
            #invoice = wb.Sheets(wb.Sheets.Count)
            invoice = wb.Sheets['DOC (2)']
            invoice.Name = 'INVOICE'

            #packingList.row_dimensions.group(55,250, hidden=True) #.delete_rows(55,2500)
            for r in range(1,packingList.UsedRange.Rows.Count):
                for c in range(1,packingList.UsedRange.Columns.Count):
                    s = packingList.Cells(r,c).Value
                    try:
                        if s and ("PACKING" in s and "LIST" in s):
                            if r > 1: packingList.Rows("1:"+str(r-1)).Hidden=True
                            packingList.Rows(str(r+57)+":"+str(packingList.UsedRange.Rows.Count)).Hidden=True
                            break
                    except Exception as e:
                        #print(e)
                        pass
                        
            for r in range(1,invoice.UsedRange.Rows.Count):
                for c in range(1,invoice.UsedRange.Columns.Count):
                    s = invoice.Cells(r,c).Value
                    try:
                        if s and ("COMMERCIAL" in s and "INVOICE" in s):  #"COMMERCIAL INVOICE" in s: 
                            #invoice.delete_rows(1,r-1)
                            if r > 1: invoice.Rows("1:"+str(r-1)).Hidden=True                      
                            invoice.Rows(str(r+57)+":"+str(packingList.UsedRange.Rows.Count)).Hidden=True
                            break
                    except Exception as e:
                        #print(e)
                        pass   
                    
            wb.Save()
            wb.Close()
        except Exception as e:
            print(e)
            pass                      

            
    #excel.Quit()


 
parse_sheet_doc_usingWin32Com(os.path.abspath(r".\InvoicesDirectory"))

# =============================================================================
def xlsx2pdf(path):
    filenames = glob.glob(path + "\*.xlsx")
    filenames = [f for f in filenames if "~" not in f] # remove temporary files
    #excel = win32.gencache.EnsureDispatch('Excel.Application') = client.Dispatch("Excel.Application")
        
    ##excel = win32.gencache.EnsureDispatch('Excel.Application') = win32.Dispatch("Excel.Application")
    
    #excel = win32.gencache.EnsureDispatch('Excel.Application')
    #excel.Visible = False
    for file in filenames:
        print("xlsx2pdf: ", file)
        try:
            books = excel.Workbooks.Open(file)
            books.D
            is_com_error = False
        except Exception as e: # Handle com_error: -2147352567 'Open method of Workbooks class failed' 'xlmain11.chm'
            is_com_error = True
            print(e)
            books = excel.Workbooks.Open(file, CorruptLoad=1)
            excel.DisplayAlerts = 0
            excel.SendKeys("{Enter}",Wait=1)    
        
        try:
            wsInvoice = books.Worksheets['INVOICE']
            wsPackageList = books.Worksheets['PACKING LIST']
        except:
            try:
                wsInvoice = books.Worksheets['INVOICE ']
                wsPackageList = books.Worksheets['PACKING LIST ']
            except:
                try:
                    wsInvoice = books.Worksheets['Invoice ']
                    wsPackageList = books.Worksheets['PList']
                except:
                    wsInvoice = books.Worksheets['INV CIP']
                    wsPackageList = books.Worksheets['PACKING LIST']
                    
        #### Some worksheet's formular not running ==> So we have to run them ###
        wsInvoice.EnableCalculation = True
        wsInvoice.Calculate()
        wsInvoice.PageSetup.Zoom = False
        wsInvoice.PageSetup.FitToPagesTall = 1
        wsInvoice.PageSetup.FitToPagesWide = 1
        wsInvoice.PageSetup.PrintGridlines = False
        
        wsPackageList.EnableCalculation = True
        wsPackageList.Calculate()
        wsPackageList.PageSetup.Zoom = False
        wsPackageList.PageSetup.FitToPagesTall = 1
        wsPackageList.PageSetup.FitToPagesWide = 1
        wsPackageList.PageSetup.PrintGridlines = False
        
        # Loop through all cells in all worksheets
        for sheet in [wsInvoice, wsPackageList]:
            unhinddenIndex = 1
            for r in range(1, sheet.UsedRange.Rows.Count):
                if sheet.Rows[r].Hidden == False:
                    unhinddenIndex = r
                    break
            for r in range(unhinddenIndex+25,unhinddenIndex+55): #range(1, sheet.UsedRange.Rows.Count+1):
                for c in range(1, sheet.UsedRange.Columns.Count):
                    try:
                        if 'H_C' in sheet.Cells(r,c).Font.Name:
                            sheet.Cells(r,c).Font.Name =  'Arial' #= sheet.cell(r,c).style.copy(font= 'Arial')
                        s = sheet.Cells(r,c).Value

                        if s != None and "/////////////////////////////////////////////////////////////" in s: 
                            sheet.Cells(r,c).Value = ''
                        elif  s != None and "AMOUNT" in s and sheet.Columns(c).ColumnWidth < 17:
                            sheet.Columns(c).ColumnWidth = 17
                    except Exception as e:
                        pass
                    
        if not is_com_error:
            books.Save()
        
        ########################################################################
        #time.sleep(5)
        #ws.Range("A1", "P100").HorizontalAlignment = 2  #align LEFT 
        #ws.Range("A1", "P80").HorizontalAlignment = 1 #align RIGHT
        try:
            wsInvoice.SaveAs(file+'INVOICE.pdf', FileFormat=57)
            wsPackageList.SaveAs(file+'PACKING LIST.pdf', FileFormat=57)
            #books.Save()
            books.Close()
        except Exception as e:
            print("Failed to convert xlsx to pdf")
            print(str(e))
        #finally:
            
    #excel.Quit()
    


xlsx2pdf(os.path.abspath(r".\InvoicesDirectory"))


#======================Convert .pdf to .png ========================
from pdf2image import convert_from_path
filenames = glob.glob('.\InvoicesDirectory' + "\*.pdf")
for file in filenames:
    images = convert_from_path(file,poppler_path=r'.\poppler-22.01.0\Library\bin')
    images[0].save(file+'.png', "PNG")
    
#Close Excel application
excel.Application.Quit()
excel.Quit()

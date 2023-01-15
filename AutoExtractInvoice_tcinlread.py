# -*- coding: utf-8 -*-
"""
Created on Wed Jul 20 17:12:40 2022

@author: samsung
"""


import numpy as np
import re
import pandas as pd
from transformers import LayoutLMv2FeatureExtractor, LayoutLMv2TokenizerFast, LayoutLMv2Tokenizer, LayoutLMv2ForTokenClassification
from ApplyLayoutLMv2_Graph_TableNet_Util import *
from py2neo import Graph
from TableDetection import TablePredict
import albumentations as album
from dateutil.parser import parse
from albumentations.pytorch.transforms import ToTensorV2

switcherTcInlRead = {
        1: "Shipper",
        2: "Invoice No",
        3: "Invoice Date",
        4: "Arrival",
        5: "Description of Goods",
        6: "N.W",
        7: "G.W",
        8: "UNIT",
        9: "PackingMethod",
        10: "Currency",
        11: "Amount",
        12: "Container \ Seal No",
        13: "Departure",
        14: "Carrier",
        15: "Sailing on About",
        16: "Notify Party",
        17: "UnitPrice",
        18: "Consignee"
}

def question2header_func(df):
    t=None
    t2=None
    for index, row in df.iterrows():
        quest = row['question'].lower()
        if 'date' in quest and 'invoice' in quest and len(row['answer']) > 5:
            try: #dateutil.parser.
                dt = parse( row['answer'])
                df.at[index,'answer'] = pd.to_datetime(dt.strftime('%d-%b-%Y') )#.strftime('%d/%m/%Y')
                df.at[index,'question2header'] = "Invoice Date"
                df.at[index,'sort'] = 3
            except:
                df.at[index,'question2header'] = "Invoice No"
                df.at[index,'sort'] = 2
        elif 'date' in quest and len(row['answer']) > 5:
            try: #dateutil.parser.
                dt = parse( row['answer'])
                df.at[index,'answer'] = pd.to_datetime(dt.strftime('%d-%b-%Y'))
                df.at[index,'question2header'] = "Invoice Date"
                df.at[index,'sort'] = 3
            except:                     
                df.at[index,'question2header'] = "Invoice No"
                df.at[index,'sort'] = 2
        elif ('contract' in quest or 'invoice' in quest) and 'no' in quest:
            df.at[index,'question2header'] = "Invoice No"
            df.at[index,'sort'] = 2
        elif 'port' in quest and 'load' in quest:
            df.at[index,'question2header'] = "Arrival"
            df.at[index,'sort'] = 4
        elif 'final' in quest and 'dest' in quest:
            df.at[index,'question2header'] = "Departure"
            df.at[index,'sort'] = 13
        elif 'carrier' in quest:
            df.at[index,'question2header'] = "Carrier"
            df.at[index,'sort'] = 14
        elif re.search(r'sai[a-z]ing', quest) is not None: # 'sailing' in quest:
            try: #dateutil.parser
                dt = parse( row['answer'])
                df.at[index,'answer'] = pd.to_datetime(dt.strftime('%d-%b-%Y'))
                df.at[index,'question2header'] = "Sailing on About"
                df.at[index,'sort'] = 15
            except:                    
                df.at[index,'question2header'] = ""
        elif 'notify' in quest and 'party' in quest:
            df.at[index,'question2header'] = "Notify Party"
            df.at[index,'sort'] = 16
        elif 'risk' in quest and 'messrs' in quest:
            df.at[index,'question2header'] = "Consignee"
            df.at[index,'sort'] = 18
        elif 'shipper' in quest or quest == 'to:':
            df.at[index,'question2header'] = "Shipper"
            df.at[index,'sort'] = 1
        elif 'amount' in quest or 'total' in quest or 'cfr value' in quest:
            df.at[index,'answer'] = row['answer'].replace(" ","")
            df.at[index,'question2header'] = "Amount"
            df.at[index,'sort'] = 11
        elif 'unit' in quest and 'price' in quest:
            df.at[index,'question2header'] = "UnitPrice"
            df.at[index,'sort'] = 17
        elif ('gross' in quest or 'gro' in quest) and ('weight' in quest or 'w' in quest or '-' in quest):
            df.at[index,'answer'] = re.sub("[() ]","",row['answer'])
            df.at[index,'question2header'] = "G.W"
            df.at[index,'sort'] = 7
        elif 'net' in quest and ('weight' in quest or 'w' in quest or '-' in quest):
            df.at[index,'answer'] = re.sub("[() ]","",row['answer'])
            df.at[index,'question2header'] = "N.W"
            df.at[index,'sort'] = 6
# =============================================================================
#             #some time: Net W'T --- 67PACKAGE
#             if re.match(r"^[0-9]+[A-Za-z]+",row['answer']):   #format numberPackingMethod ex: 68Pallet(s), 87Roll(s)
#                 t = pd.DataFrame({'question':[ "UNIT", "PackingMethod"],
#                                   'question2header':["UNIT", "PackingMethod"], 
#                                   'answer':[re.findall(r'\d+', row['answer'])[0], re.findall(r'[A-Za-z]+', row['answer'])[0]], 
#                                   'sort':[8, 9]})
#                 df = df.append(t)
# =============================================================================
        elif 'description' in quest and 'goods' in quest:
            df.at[index,'question2header'] = "Description of Goods"
            if '(PLACE OF TERMS OF PRICE)' in row['answer'] :
                df.at[index,'question2header'] += "(TOBE DELETED)"
                df.at[index,'question'] += "(TOBE DELETED)"
            else:
                df.at[index,'sort'] = 5
        elif 'container' in quest:
            df.at[index,'question2header'] = "Container \ Seal No"
            df.at[index,'sort'] = 12
        elif 'quantity' in quest:
            #pattern = re.compile("^[0-9]+[A-Za-z(A-Za-z)]+")
            if re.match(r"^[0-9]+[A-Za-z]+",row['answer']) \
                or re.match(r"^[0-9]+ [A-Za-z]+",row['answer']):   #format numberPackingMethod ex: 68Pallet(s), 87Roll(s)
                unit = float(re.findall(r'\d+', row['answer'])[0])
                if t is None or t['answer'][0] < unit:
                    t = pd.DataFrame({'question':[ "UNIT", "PackingMethod"],
                                      'question2header':["UNIT", "PackingMethod"], 
                                      'answer':[unit, re.findall(r'[A-Za-z]+', row['answer'])[0]], 
                                      'sort':[8, 9]})
                #df = df.append(t) #cannot append in for loop
        elif 'roll' in quest or 'pallet' in quest or 'package' in quest :
            unit = None
            try:
                unit = float(row['answer'].split('-')[-1])
            except:
                unit = row['answer'].split('-')[-1]
            if t2 is None:
                t2 = pd.DataFrame({'question':[ "UNIT", "PackingMethod"],
                                  'question2header':["UNIT", "PackingMethod"], 
                                  'answer':[unit, quest.split(' ')[0].upper() ], 
                                  'sort':[8, 9]})       
            else:
                t2 = t2.append( pd.DataFrame({'question':[ "UNIT", "PackingMethod"],
                                  'question2header':["UNIT", "PackingMethod"], 
                                  'answer':[unit, quest.split(' ')[0].upper() ], 
                                  'sort':[8, 9]}) )
    if t is not None:
        return df.append(t)
    if t2 is not None:
        return df.append(t2)
    return df


def extQuestionAnswerGraphFunc(graph, file=None):
    
    ################# CONNECT "Container No." to Answer #########################
    #CASE: file "TCUSCF-2207-03 1(NYLON 018)GY.xls" like below:
    #	          CONTAINER NO.		
    #   BEAU5278373 / VN1626314A	
    tx = graph.begin()    
    tx.evaluate('''
                match (q1:QUESTION), (q2:QUESTION), (a:ANSWER)
                where ((q1:QUESTION)-[]->(q2:QUESTION)) //
                    and ANY( x in ['CONTAINER'] where q1.word contains x)
                    and not(()-[]->(a:ANSWER)) //start word in answer string
                    and abs(a.y0 - q1.y1) < 35
                    and abs(a.x0 - q1.x0) < $width/3.5
                create (q2)-[:RELATED {y:abs(a.y0 - q1.y1), x:abs(a.x1 - q1.x0)}]->(a);
                ''', parameters = {'width': 1700})
    graph.commit(tx)
    
    #CASE: Summarize table (Description of Goods)
    focusHigh = 200
    tx = graph.begin()
    tx.evaluate('''
                with ['Party', 'Messrs'] as notifyParty_Consignee 
                match (qstart:QUESTION), (qend:QUESTION), (a:ANSWER)
                where not(()-[]->(qstart:QUESTION)) //start word in A question string
                    and not ((qend:QUESTION)-[]->(:QUESTION)) //end word in A question string
                    and ((qstart:QUESTION)-[*0..]->(qend:QUESTION)) //a path from start 2 end (qstart->qend)
                    and not((:ANSWER)-[]->(a:ANSWER)) //start word in answer string
                    and not((:QUESTION)-[]->(a:ANSWER)) 
                    and a.y0 - qstart.y1 > 0 and a.y0 - qstart.y1 < $focusHigh //  280 không lấy xuống quá nhiều
                    and abs(a.x0 - qstart.x0) < 60 //45  --date modify: 20220725
                    and any (word in notifyParty_Consignee where qend.word contains word)
                    //and not any (word in ['SAVANNAH', 'VIETNAM'] where a.word contains word)
                    //and not exists ( OPTIONAL MATCH  (ansAboveA:ANSWER)
                    //                    where abs(ansAboveA.x0 - a.x0) < 5  and  0 < a.y0-ansAboveA.y0 )
                create (qend)-[:RELATED {y:abs(a.y0 - qend.y1), x:abs(a.x1 - qend.x0)}]->(a);
                ''', parameters = {'focusHigh': focusHigh})
    graph.commit(tx)
        
def trySplitKGS(value):
    try:
        if 'KGS' in value:
            return float( value.split('KGS',1)[0].strip().replace(',','') )
        elif 'KG' in value:
            return float( value.split('KG',1)[0].strip().replace(',','') )
    except:
        return -1    
###################################### MAIN ###################################
feature_extractor = LayoutLMv2FeatureExtractor.from_pretrained(r".\savedmodel")#("microsoft/layoutlmv2-base-uncased")
tokenizer = LayoutLMv2TokenizerFast.from_pretrained(r".\savedmodel")#("microsoft/layoutlmv2-base-uncased")
model = LayoutLMv2ForTokenClassification.from_pretrained(r".\savedmodel")

#uri = "neo4j+s://dff0ff04.databases.neo4j.io"
uri = "neo4j://localhost:7687"
user = "neo4j"
#password = "CJY__L21_jWwCzLqipfOM0WrAoPOH7LEomMCfN2GFN0"
password = "pa@ss"
graph = Graph(uri, auth=(user, password))

transforms = album.Compose([
    album.Resize(896, 896, always_apply=True),
    album.Normalize(),
    ToTensorV2()
])
tablePred = TablePredict(r'.\savedmodel\best_model.ckpt', transforms)

lstHasToBeQuestion=["Goods","Good","goods","good","CFR","TOTAL","Total","cfr", "Sailing","Shipper"]
lstHasToBeAnswer=["HYOSUNG", "VIETNAM", "USA","CO.,LTD.","CO.,LTD","CO.LTD","TEL", "Tel","ATTN", "FAX","Fax", '[\d]+ROLL', '[\d]+ ROLL', '[\d]+ PAL', '[\d]+PAL']

output_dictionary_png_df = {}
invoice_info_extract_from_png(feature_extractor, tokenizer, model, graph, os.path.abspath(r".\InvoicesDirectory"), question2header_func, 
                              switcher = switcherTcInlRead, tablePred=tablePred, 
                              out_collection=output_dictionary_png_df,
                              cut_image=["left","bottom"],
                              lstHasToBeQuestion=lstHasToBeQuestion,
                              lstHasToBeAnswer=lstHasToBeAnswer,
                              extQuestionAnswerGraphFunc=extQuestionAnswerGraphFunc)



      
################################### WRITE OUTPUT ##############################
from openpyxl import load_workbook, Workbook

book = Workbook() #load_workbook('./output.xlsx')
writer = pd.ExcelWriter(r'./InvoicesDirectory/output.xlsx', engine='openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
for df_name, df in output_dictionary_png_df.items():
    df.to_excel(writer, sheet_name=df_name)

writer.save()
writer.close()
print("Finish export output.xls")

################################# WRITE UPLOAD FILE ############################
output_dictionaryFinal = {}
for k, v in output_dictionary_png_df.items():  
    print("Process output_dictionary_png_df: ",k)
    ############## DO Merge step of INVOICE and PACKINGLIST#####################
    file = k.split('.',1)[0]
    temp = v[['question2header', 'answer','sort']]
    temp.replace("", float("NaN"), inplace=True)
    temp.dropna(subset=['question2header'], inplace=True)
    temp.reset_index(drop=True, inplace=True)
    if file in output_dictionaryFinal:
        tempV = output_dictionaryFinal[file]

        tempV_NotifyConsignee =  tempV.loc[tempV['question2header'].isin(['Notify Party','Consignee'])]
        tempV = tempV.loc[~tempV['question2header'].isin(['Notify Party','Consignee'])]
        tempV_NotifyConsignee = (tempV_NotifyConsignee.dropna(subset=['answer'])
                        .assign(something=lambda x: x['answer'].str.len())
                        .sort_values(['question2header','something'], ascending=[True, False])
                        .groupby('question2header', as_index=False)
                        .head(1))
        # tempV_NotifyConsignee['answer'] = tempV_NotifyConsignee.groupby(['question2header'])['answer'].transform(lambda x : '\n'.join(x))
        # tempV_NotifyConsignee = tempV_NotifyConsignee.drop_duplicates()
        
        temp_NotifyConsignee =  temp.loc[temp['question2header'].isin(['Notify Party','Consignee'])]
        temp = temp.loc[~temp['question2header'].isin(['Notify Party','Consignee'])]
        temp_NotifyConsignee = (temp_NotifyConsignee.dropna(subset=['answer'])
                        .assign(something=lambda x: x['answer'].str.len())
                        .sort_values(['question2header','something'], ascending=[True, False])
                        .groupby('question2header', as_index=False)
                        .head(1))
        # temp_NotifyConsignee['answer'] = temp_NotifyConsignee.groupby(['question2header'])['answer'].transform(lambda x : '\n'.join(x))
        # temp_NotifyConsignee = temp_NotifyConsignee.drop_duplicates()
        
        notifyConsigneePD = pd.merge(temp_NotifyConsignee, tempV_NotifyConsignee, how='outer', left_on = ['question2header','sort'], right_on = ['question2header','sort']).drop_duplicates()
        
        notifyConsigneePD['answer'] = ""
        for index, row in notifyConsigneePD.iterrows():
            try:
                notifyConsigneePD.at[index,'answer'] = row['answer_x'] if len(row['answer_x']) > len(row['answer_y']) else row['answer_y']
            except:
                notifyConsigneePD.at[index,'answer'] = row['answer_x'] if checkNanValue(row['answer_y']) else row['answer_y']
                continue
        notifyConsigneePD.drop_duplicates(inplace=True)
            
        finalPD = pd.merge(temp, tempV, how='outer', left_on = ['question2header','sort'], right_on = ['question2header','sort']).drop_duplicates()
        # lay data theo condtion
        finalPD['answer'] = ""
        currency = ""
        for index, row in finalPD.iterrows():
            quest = row['question2header']
            if quest in ['Departure', 'Arrival']:
                try:
                    finalPD.at[index,'answer'] = row['answer_x'] if len(row['answer_x']) > len(row['answer_y']) else row['answer_y']
                except:
                    finalPD.at[index,'answer'] = row['answer_x'] if pd.notna(row['answer_x']) else row['answer_y']
            elif quest in ['Carrier', 'Shipper']: #this condition is opposite with 'Invoice No'
                try:
                    finalPD.at[index,'answer'] = row['answer_x'] if len(row['answer_x']) > len(row['answer_y']) else row['answer_y']
                except:
                    finalPD.at[index,'answer'] = row['answer_x'] if pd.notna(row['answer_x']) else row['answer_y']
            elif 'Invoice No' in quest:
                ansx = (row['answer_y'][1:] if row['answer_y'][0] == 'I' else  row['answer_y']) if pd.isna(row['answer_x']) else (row['answer_x'][1:] if row['answer_x'][0] == 'I' else row['answer_x'])
                ansy = (row['answer_x'][1:] if row['answer_x'][0] == 'I' else  row['answer_x']) if pd.isna(row['answer_y']) else (row['answer_y'][1:] if row['answer_y'][0] == 'I' else row['answer_y'])
                finalPD.at[index,'answer'] = ansx if ansx == ansy else ansy                
                # maybeInvoiceNo = file.split(' ',1)[0]
                # if maybeInvoiceNo != finalPD.at[index,'answer']:#force the right value, case: file name of *.xls always start with InvoiceNO 
                #     finalPD.at[index,'answer'] = maybeInvoiceNo 
                if  ' 'in  finalPD.at[index,'answer']:
                    finalPD.at[index,'answer'] = finalPD.at[index,'answer'].split(' ')[0]
                elif len(finalPD.at[index,'answer']) < 8:
                    finalPD.at[index,'question2header'] += "(TOBE DELETED)"
            elif quest in ['G.W', 'N.W', 'Notify Party', 'Consignee', 'Sailing on About', 'UnitPrice', 'Invoice Date', 'Container \ Seal No', 'Description of Goods','UNIT','PackingMethod']:
                finalPD.at[index,'answer'] = row['answer_y']  if pd.isna(row['answer_x']) else row['answer_x']
            elif quest in ['Amount']:
                ansx = "" if pd.isna(row['answer_x']) else row['answer_x']
                ansy = "" if pd.isna(row['answer_y']) else row['answer_y']
                if Price.fromstring(ansx).currency in ['USD', 'EUR', 'JPY', 'VND', 'LB', '$','£','¥','€']:
                    finalPD.at[index,'answer'] = float(Price.fromstring(ansx).amount)
                    currency = Price.fromstring(ansx).currency
                elif Price.fromstring(ansy).currency in ['USD', 'EUR', 'JPY', 'VND', 'LB', '$','£','¥','€']:
                    finalPD.at[index,'answer'] = float(Price.fromstring(ansy).amount)
                    currency = Price.fromstring(ansy).currency
        ####################### REFINE FINAL ##################################
        filter_TODELETE = finalPD['question2header'].str.contains('TOBE DELETED') == False
        finalPD = finalPD.where(filter_TODELETE)
        finalPD = finalPD.append(notifyConsigneePD)
            
        finalPD['answer'].replace("", float("NaN"), inplace=True)
        finalPD.dropna(subset=['answer'], inplace=True)
        finalPD_addCurrency = finalPD[['question2header', 'answer', 'sort']].drop_duplicates()
        finalPD_addCurrency = finalPD_addCurrency.append(pd.DataFrame([["Currency", currency, 10]], columns=['question2header','answer','sort']))
        finalPD_addCurrency.reset_index(drop=True, inplace=True)
        
        invoieNoDf = finalPD_addCurrency.loc[finalPD_addCurrency['question2header'].isin(['Invoice No'])]
        invoieNoDf['answer'] = invoieNoDf.groupby(['question2header'])['answer'].transform(lambda x : '\n'.join(x)).drop_duplicates()
        invoieNoDf = invoieNoDf.drop_duplicates()
        invoieNoDf.dropna(inplace=True)
        
        arrivalDF = finalPD_addCurrency.loc[finalPD_addCurrency['question2header'].isin(['Arrival'])]
        arrivalDF = arrivalDF[0:1]
        
        temDF = finalPD_addCurrency.loc[finalPD_addCurrency['question2header'].isin(['Departure'])]
        lengths = temDF["answer"].str.len()
        argmax = np.where(lengths == lengths.max())[0]        
        departDF = temDF.iloc[argmax][-1:]
        
        temDF = finalPD_addCurrency.loc[finalPD_addCurrency['question2header'].isin(['Carrier'])]
        lengths = temDF["answer"].str.len()
        argmax = np.where(lengths == lengths.max())[0]        
        carrierDF = temDF.iloc[argmax][-1:]
        
        temDF =  finalPD_addCurrency.loc[(finalPD_addCurrency['question2header'].isin(['Shipper'])) & (finalPD_addCurrency["answer"].str.len()>10) ]
        lengths = temDF["answer"].str.len()
        argmin = np.where(lengths == lengths.min())[0]
        shipperDF = temDF.iloc[argmin]
        
        # temDF = finalPD_addCurrency.loc[finalPD_addCurrency['question2header'].isin(['Notify Party'])]
        # lengths = temDF["answer"].str.len()
        # argmax = np.where(lengths == lengths.max())[0]        
        # notifyPartyDF = temDF.iloc[argmax][-1:]
        
        #notifyPartyDF = finalPD_addCurrency.loc[finalPD_addCurrency['question2header'].isin(['Notify Party', 'Consignee'])]
        #notifyPartyDF['answer'] = notifyPartyDF.groupby(['question2header'])['answer'].transform(lambda x : '\n'.join(x))
        #notifyPartyDF = notifyPartyDF.drop_duplicates()

        filterGross = finalPD_addCurrency['question2header'].isin(['G.W'])
        filterNet = finalPD_addCurrency['question2header'].isin(['N.W'])
        filter2 = finalPD_addCurrency['answer'].str.contains('KGS') #only get KGS, not other ~LBS ~Ib
        filter3 = finalPD_addCurrency['answer'].str.contains('LBS') == False
        filter4 = finalPD_addCurrency['answer'].str.contains('LB') == False
        filter5 = finalPD_addCurrency['answer'].str.contains('Ib') == False
        #grossNetDF = finalPD_addCurrency.loc[finalPD_addCurrency['question2header'].isin(['G.W', 'N.W'])]
        grossDF = finalPD_addCurrency.where(filterGross & (filter3 | filter4 | filter5)).dropna()
        netDF = finalPD_addCurrency.where(filterNet & (filter3 | filter4 | filter5)).dropna()
        # concatenate the string
        try:
            grossDF['answer'] = grossDF['answer'].apply(lambda x: trySplitKGS(x) )
            netDF['answer'] = netDF['answer'].apply(lambda x: trySplitKGS(x) )
            grossMax = grossDF.groupby(['question2header'])['answer'].transform(max).iloc[0]
            netMax = netDF.groupby(['question2header'])['answer'].transform(max).iloc[0]
            if netMax == grossMax:
                netDF = netDF.loc[netDF['answer'] != netMax]
                netMax = netDF.groupby(['question2header'])['answer'].transform(max).iloc[0]
                
            grossDF['answer'] = grossMax 
            grossDF = grossDF.drop_duplicates()
            netDF['answer'] = netMax 
            netDF = netDF.drop_duplicates()
        except:
            grossDF['answer'] = ""
            netDF['answer'] = ""

        
        amountDF = finalPD_addCurrency.loc[finalPD_addCurrency['question2header'].isin(['Amount'])]
        try:
            column = pd.to_numeric(amountDF['answer'])
            idxmax = column.idxmax()
            amountDF = amountDF.loc[idxmax]
        except Exception as e:
            amountDF = amountDF[-1:]
            pass
        
        descriptionOfGoodsDF = finalPD_addCurrency.loc[finalPD_addCurrency['question2header'].isin(['Description of Goods'])]
        #descriptionOfGoodsDF = descriptionOfGoodsDF[-1:]
        #descriptionOfGoodsDF = descriptionOfGoodsDF[0:-1] #Delete the thired row, only accept the first 2 rows
        descriptionOfGoodsDF = descriptionOfGoodsDF[descriptionOfGoodsDF["answer"].str.len() >= 15]
        descriptionOfGoodsDF['answer'] = descriptionOfGoodsDF.groupby(['question2header'])['answer'].transform(lambda x : '\n'.join(x))
        descriptionOfGoodsDF = descriptionOfGoodsDF.drop_duplicates()
        
        unitDF = finalPD_addCurrency.loc[finalPD_addCurrency['question2header'].isin(['UNIT'])]
        try:
            column = pd.to_numeric(unitDF['answer'])
            idxmax = column.idxmax()
            unitDF = unitDF.loc[idxmax]
        except Exception as e:
            unitDF = unitDF[-1:]
            pass
        
        packingMethodDF = finalPD_addCurrency.loc[finalPD_addCurrency['question2header'].isin(['PackingMethod'])]        
        packingMethodDF = packingMethodDF.drop(packingMethodDF[packingMethodDF['answer'].str.isdigit()].index)
        packingMethodDF = packingMethodDF[0:1]
        
        unitPriceDF = finalPD_addCurrency.loc[finalPD_addCurrency['question2header'].isin(['UnitPrice'])]
        unitPriceDF = unitPriceDF[unitPriceDF['answer'].str.contains('\d')]
        unitPriceDF['answer'] = unitPriceDF.groupby(['question2header'])['answer'].transform(lambda x : ', '.join(x))
        unitPriceDF = unitPriceDF.drop_duplicates()
        
        temDF = finalPD_addCurrency.loc[finalPD_addCurrency['question2header'].isin(['Container \ Seal No'])]
        lengths = temDF["answer"].str.len()
        argmax = np.where(lengths == lengths.max())[0]        
        containerNoSealNoDF = temDF.iloc[argmax]
        

        restDF =  finalPD_addCurrency.loc[~finalPD_addCurrency['question2header'].isin(['Invoice No','Arrival','Departure','G.W','N.W','UnitPrice','Shipper',
                                                                                        'Description of Goods','Carrier','PackingMethod',
                                                                                        'UNIT', 'Amount','Container \ Seal No', 'Notify Party','Consignee'])]
        refinedFinalDF = invoieNoDf.append(arrivalDF).append(departDF).append(grossDF).append(netDF).append(descriptionOfGoodsDF).append(shipperDF) \
                                .append(amountDF).append(unitPriceDF).append(carrierDF).append(unitDF).append(packingMethodDF) \
                                .append(containerNoSealNoDF).append(notifyConsigneePD).append(restDF)
        missData = findMissingData(refinedFinalDF['sort'].tolist(), switcher=switcherTcInlRead)
        #add the missing data
        for miss in missData:
            refinedFinalDF = refinedFinalDF.append(pd.DataFrame([[numbers_to_header(miss, switcher=switcherTcInlRead), "", miss]], columns=['question2header','answer','sort']))
        refinedFinalDF = refinedFinalDF.append(pd.DataFrame([["File", file, 0]], columns=['question2header','answer','sort']))
        refinedFinalDF = refinedFinalDF.sort_values(by=['sort'])
        refinedFinalDF.reset_index(drop=True, inplace=True)
        output_dictionaryFinal[file] = refinedFinalDF
    else:
        output_dictionaryFinal[file] = temp
     


############################################## EXPORT FINAL RESULT TO *.xlsx ###############################################
exportDF = None
for df_name, df in output_dictionaryFinal.items():
    print("Export file: ",df_name)
    if exportDF is None:
        t = df[['question2header', 'answer']]
        t.set_index('question2header', inplace=True)
        exportDF = t.T
    else:
        t = df[['question2header', 'answer']]
        t.set_index('question2header', inplace=True)
        exportDF = exportDF.append(t.T)

writer = pd.ExcelWriter(r'.\InvoicesDirectory\outputFinal_tcinlread.xlsx',
                        engine='xlsxwriter',
                        datetime_format='dd/MMM/yyyy',
                        date_format='dd/MMM/yyyy')
#exportDF.to_excel(r'.\InvoicesDirectory\outputFinal_tcinlread.xls')

exportDF.to_excel(writer, sheet_name='Sheet1')
workbook  = writer.book
worksheet = writer.sheets['Sheet1']
# Get the dimensions of the dataframe.
(max_row, max_col) = df.shape
# Set the column widths, to make the dates clearer.
worksheet.set_column(1, max_col, 20)
writer.save()
writer.close()
        

print("Finish export outputFinal_tcinlread.xlsx")